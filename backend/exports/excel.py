from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

from backend.services.analytics import build_options_analytics, build_portfolio_analytics, build_risk_analytics


def build_workbook_bytes(months: int = 6, risk_free_rate: float = 0.07) -> bytes:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Options"

    options_data = build_options_analytics(months=months, risk_free_rate=risk_free_rate)
    portfolio_data = build_portfolio_analytics(months=months, risk_free_rate=risk_free_rate)
    risk_data = build_risk_analytics(months=months)

    summary_sheet.append(
        [
            "Symbol",
            "Contract",
            "Strike",
            "Expiry",
            "Market Price",
            "Market Source",
            "Market IV",
            "Historical Vol",
            "GARCH IV",
            "BSM Hist",
            "BSM GARCH",
            "BSM Market IV",
            "Hist Delta",
            "Hist Gamma",
            "Hist Vega",
            "Hist Theta",
            "Hist Rho",
            "GARCH Delta",
            "GARCH Gamma",
            "GARCH Vega",
            "GARCH Theta",
            "GARCH Rho",
            "Market Delta",
            "Market Gamma",
            "Market Vega",
            "Market Theta",
            "Market Rho",
        ]
    )
    for cell in summary_sheet[1]:
        cell.font = Font(bold=True)
    for symbol_block in options_data["symbols"]:
        for contract in symbol_block["contracts"]:
            market_greeks = contract.get("greeks_market_vol") or {}
            summary_sheet.append(
                [
                    symbol_block["symbol"],
                    contract["label"],
                    contract["strike"],
                    contract["expiry_date"],
                    contract["market_price"],
                    contract["market_price_source"],
                    contract["market_implied_volatility"],
                    contract["historical_volatility"],
                    contract["garch_volatility"],
                    contract["bsm_historical_vol_price"],
                    contract["bsm_garch_vol_price"],
                    contract.get("bsm_market_vol_price"),
                    contract["greeks_historical_vol"]["delta"],
                    contract["greeks_historical_vol"]["gamma"],
                    contract["greeks_historical_vol"]["vega"],
                    contract["greeks_historical_vol"]["theta"],
                    contract["greeks_historical_vol"]["rho"],
                    contract["greeks_garch_vol"]["delta"],
                    contract["greeks_garch_vol"]["gamma"],
                    contract["greeks_garch_vol"]["vega"],
                    contract["greeks_garch_vol"]["theta"],
                    contract["greeks_garch_vol"]["rho"],
                    market_greeks.get("delta"),
                    market_greeks.get("gamma"),
                    market_greeks.get("vega"),
                    market_greeks.get("theta"),
                    market_greeks.get("rho"),
                ]
            )

    portfolio_sheet = workbook.create_sheet("Portfolio")
    portfolio_sheet.append(["Symbol", "Bucket", "Delta", "Gamma", "Vega", "Hedge Shares", "Adjusted Hedge"])
    for cell in portfolio_sheet[1]:
        cell.font = Font(bold=True)
    for row in portfolio_data["portfolios"]:
        portfolio_sheet.append(
            [
                row["symbol"],
                row["bucket"],
                row["portfolio_greeks"]["delta"],
                row["portfolio_greeks"]["gamma"],
                row["portfolio_greeks"]["vega"],
                row["hedge"]["raw_shares"],
                row["hedge"]["liquidity_adjusted_shares"],
            ]
        )

    risk_sheet = workbook.create_sheet("Risk")
    risk_sheet.append(["Symbol", "Bucket", "Regime", "VaR 95 Parametric", "VaR 99 Parametric", "VaR 95 GARCH", "VaR 99 GARCH"])
    for cell in risk_sheet[1]:
        cell.font = Font(bold=True)
    for row in risk_data["comparison_table"]:
        risk_sheet.append(
            [
                row["symbol"],
                row["bucket"],
                row["regime"],
                row["var_95_parametric"],
                row["var_99_parametric"],
                row["var_95_garch"],
                row["var_99_garch"],
            ]
        )

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
